# -*- coding: utf-8 -*-
"""
Created on Mon Dec 17 18:56:32 2018

@author: Michael
"""

import openpyxl as ox
import csv
import numpy as np
import pytesseract as pt
import cv2
num_files=3 #change this value of number of files
sample="A4"# change this based on the sample
width = []
for num in range(num_files):  #filter through picture using OpenCV
    file_name1= sample+"_NA_" + str(num+1) + ".jpg"
    #save_name = "tearanalysis_" + sample + "_NA_" + str(num+1) + ".jpg"
    raw = cv2.imread(file_name1)
    h, w = raw.shape[:2]# initially cropping image 
    img= raw[100:h,0:w]
    #cv2.imshow("img",img)
    
    rimg = cv2.resize(img, (0,0), fx=0.5, fy=0.5) #resizing
    #cv2.imshow("resized",rimg)

    kernel = np.ones((15,15),np.uint8)
    e = cv2.erode(rimg,kernel,iterations = 2)#eroding
    #cv2.imshow("erode",e)

    gray = cv2.cvtColor(e, cv2.COLOR_BGR2GRAY)
    #cv2.imshow("gray",gray)
    ret, th = cv2.threshold(gray,5, 255, cv2.THRESH_BINARY_INV)
    #cv2.imshow("th",th)
    
    im2, cnt, h = cv2.findContours(th, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    #print(cnt[0][0][0][0],cnt[0][0][0][1])#prints coordinates
    #print(cnt[0][2][0][0],cnt[0][1][0][1])
    crop= img[cnt[0][0][0][1]*2:cnt[0][1][0][1]*2,cnt[0][0][0][0]*2:cnt[0][2][0][0]*2]
    
    #cv2.imshow("Crop im",crop)
    #cv2.waitKey(0)
    #cv2.destroyAllWindows()
    
    #im = Image.open(crop)
    text = pt.image_to_string(crop, lang = "eng")# uses Tessercact to find width from cropped img
    #print(text)
    width += [round(float("".join(text[11:-2])),1)/1000]
    
    print(width)

#%%
time = []
temp = []
strain = []
stress = []
modulus = []
length = []
gap = []
force = []
aforce = []
#%%extract data from raw CSV files and paste into template
for num in range(num_files):
    file_name2= "PCU_"+ sample+"_NA_" + str(num+1) + ".csv"
    save_name = "tearanalysis_" + sample + "_NA_" + str(num+1) + ".xlsx"
    with open(file_name2) as csvfile:
        counter=0
        r = csv.reader(csvfile, delimiter=',')
        #rowcount = sum(1 for rows in readCSV)
        for k in range(9): # count from 0 to 8
            next(r)     # and discard the rows
        for row in r:
            if row:
                time += [float(row[0])]
                temp += [float(row[1])]
                strain += [float(row[2])]
                stress += [float(row[3])]
                length +=[float(row[5])]
                gap += [float(row[6])]
                force += [float(row[7])]
                aforce +=[float(row[8])]
                if row[4]:
                     modulus += [float(row[4])]    
                else:
                    modulus += [0]
    f_array = np.array([np.array(time),np.array(temp),np.array(strain),np.array(stress),np.array(modulus),np.array(length),
                        np.array(gap),np.array(force),np.array(aforce)])
    f_array = np.transpose(f_array)
#%%       
    template = ox.load_workbook("tearanalysis_template.xlsx") 
    temp_sheet = template.get_sheet_by_name("Sheet1") 
    temp_sheet2 = template.get_sheet_by_name("Tear Analysis")
#%%            
    def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1): 
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

#%%
    pastingRange = pasteRange(1,10,len(f_array[0]),9+len(time),temp_sheet, f_array)
    temp_sheet2.cell(6,2).value = width[num]
    template.save(save_name)
    time = []
    temp = []
    strain = []
    stress = []
    modulus = []
    length = []
    gap = []
    force = []
    aforce = []
   
    
